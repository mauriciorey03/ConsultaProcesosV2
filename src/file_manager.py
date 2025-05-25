#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gestor de archivos para el sistema de consulta de procesos judiciales
Excel simplificado solo con datos
"""

import pandas as pd
import logging
import shutil
import time
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple
import json
import csv

try:
    from config import FileConfig, ProcessConfig
except ImportError:
    import sys
    from pathlib import Path
    sys.path.append(str(Path(__file__).parent))
    from config import FileConfig, ProcessConfig

try:
    from data_processor import ProcesoInfo, EstadisticasProcesamiento
except ImportError:
    import sys
    from pathlib import Path
    sys.path.append(str(Path(__file__).parent))
    from data_processor import ProcesoInfo, EstadisticasProcesamiento


logger = logging.getLogger(__name__)


class FileManagerError(Exception):
    """Excepción personalizada para errores del gestor de archivos"""
    pass


class ExcelReader:
    """Lector de archivos Excel para radicados"""
    
    def __init__(self, archivo_path: Path = None):
        """
        Inicializa el lector Excel
        
        Args:
            archivo_path: Ruta al archivo Excel (opcional, usa config por defecto)
        """
        self.archivo_path = archivo_path or FileConfig.EXCEL_INPUT_FILE
        logger.info(f"ExcelReader inicializado con archivo: {self.archivo_path}")
    
    def validar_archivo(self) -> bool:
        """
        Valida que el archivo Excel existe y es accesible
        
        Returns:
            True si el archivo es válido, False en caso contrario
        """
        try:
            if not self.archivo_path.exists():
                logger.error(f"Archivo Excel no existe: {self.archivo_path}")
                return False
            
            if not self.archivo_path.is_file():
                logger.error(f"La ruta no es un archivo: {self.archivo_path}")
                return False
            
            # Intentar leer las primeras filas para validar formato
            df = pd.read_excel(self.archivo_path, nrows=5, header=None, dtype=str)
            
            if df.empty:
                logger.error("El archivo Excel está vacío")
                return False
            
            logger.info("Archivo Excel validado correctamente")
            return True
            
        except Exception as e:
            logger.error(f"Error al validar archivo Excel: {e}")
            return False
    
    def leer_radicados(self) -> List[str]:
        """
        Lee los radicados desde el archivo Excel
        
        Returns:
            Lista de radicados encontrados
            
        Raises:
            FileManagerError: Si hay problemas al leer el archivo
        """
        try:
            if not self.validar_archivo():
                raise FileManagerError(f"Archivo Excel inválido: {self.archivo_path}")
            
            logger.info(f"Leyendo radicados desde: {self.archivo_path}")
            
            # Leer archivo Excel sin header
            df = pd.read_excel(self.archivo_path, header=None, dtype=str)
            
            radicados = []
            
            # Verificar que hay datos en la columna A
            if len(df.columns) > 0:
                # Empezar desde la fila especificada en config (default: 2, índice 1)
                start_row = FileConfig.EXCEL_START_ROW - 1  # Convertir a índice 0-based
                
                for i in range(start_row, len(df)):
                    valor = df.iloc[i, 0]  # Columna A (índice 0)
                    
                    # Procesar el valor
                    if pd.notna(valor):
                        radicado = str(valor).strip()
                        
                        if radicado and radicado != '':
                            # Validar formato básico del radicado
                            if radicado.isdigit() and len(radicado) >= ProcessConfig.MIN_RADICADO_LENGTH:
                                radicados.append(radicado)
                                logger.debug(f"Radicado válido encontrado: {radicado}")
                            else:
                                logger.warning(f"Radicado con formato inválido ignorado: {radicado}")
                        else:
                            # Celda vacía - detener lectura
                            logger.debug(f"Celda vacía encontrada en fila {i+1}, deteniendo lectura")
                            break
                    else:
                        # Celda vacía (NaN) - detener lectura
                        logger.debug(f"Celda NaN encontrada en fila {i+1}, deteniendo lectura")
                        break
            
            logger.info(f"Se encontraron {len(radicados)} radicados válidos")
            
            # Mostrar muestra de radicados para confirmación
            if radicados:
                muestra = min(5, len(radicados))
                logger.info("Primeros radicados encontrados:")
                for i, rad in enumerate(radicados[:muestra]):
                    logger.info(f"  {i+1}. {rad}")
                
                if len(radicados) > muestra:
                    logger.info(f"  ... y {len(radicados) - muestra} más")
            
            return radicados
            
        except Exception as e:
            error_msg = f"Error al leer radicados desde Excel: {e}"
            logger.error(error_msg)
            raise FileManagerError(error_msg)
    
    def obtener_info_archivo(self) -> Dict[str, Any]:
        """
        Obtiene información del archivo Excel
        
        Returns:
            Diccionario con información del archivo
        """
        try:
            if not self.archivo_path.exists():
                return {"error": "Archivo no existe"}
            
            stats = self.archivo_path.stat()
            df = pd.read_excel(self.archivo_path, header=None, dtype=str)
            
            return {
                "ruta": str(self.archivo_path),
                "tamaño_bytes": stats.st_size,
                "fecha_modificacion": datetime.fromtimestamp(stats.st_mtime).isoformat(),
                "filas_totales": len(df),
                "columnas_totales": len(df.columns),
                "primera_celda": str(df.iloc[0, 0]) if not df.empty else "Vacío"
            }
            
        except Exception as e:
            logger.error(f"Error al obtener info del archivo: {e}")
            return {"error": str(e)}


class ResultWriter:
    """Escritor de resultados - Solo Excel simplificado"""
    
    def __init__(self, output_dir: Path = None):
        """
        Inicializa el escritor de resultados
        
        Args:
            output_dir: Directorio de salida (opcional, usa config por defecto)
        """
        self.output_dir = output_dir or FileConfig.OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"ResultWriter inicializado con directorio: {self.output_dir}")
    
    def generar_nombre_archivo(self, prefijo: str, extension: str) -> str:
        """
        Genera un nombre de archivo único con timestamp
        
        Args:
            prefijo: Prefijo del nombre de archivo
            extension: Extensión del archivo (sin punto)
            
        Returns:
            Nombre de archivo único
        """
        timestamp = datetime.now().strftime(FileConfig.OUTPUT_DATETIME_FORMAT)
        return f"{prefijo}_{timestamp}.{extension}"
    
    def escribir_resultados_excel(self, procesos: List[ProcesoInfo]) -> Path:
        """
        Escribe los resultados en formato Excel simplificado
        
        Args:
            procesos: Lista de procesos procesados
            
        Returns:
            Ruta del archivo Excel creado
        """
        try:
            nombre_archivo = self.generar_nombre_archivo("consulta_procesos", "xlsx")
            ruta_archivo = self.output_dir / nombre_archivo
            
            logger.info(f"Escribiendo Excel en: {ruta_archivo}")
            
            # Preparar datos completos con las nuevas columnas
            datos_excel = []
            
            for i, proceso in enumerate(procesos, 1):
                datos_excel.append({
                    'No.': i,
                    'Radicado': proceso.radicado,
                    'Demandante': proceso.demandante,
                    'Demandado': proceso.demandado,
                    'Juzgado': proceso.juzgado,
                    'Departamento': proceso.departamento,
                    'Tipo del Proceso': proceso.tipo_proceso,
                    'Clase del Proceso': proceso.clase_proceso,
                    'Subclase del Proceso': proceso.subclase_proceso,
                    'Última Fecha de Actuación': proceso.fecha_ultima_actuacion,
                    'Última Actuación': proceso.ultima_actuacion,  # NUEVO
                    'Anotaciones': proceso.anotaciones,  # NUEVO
                    'Es Privado': 'Sí' if proceso.es_privado else 'No',
                    'Estado': proceso.status
                })
            
            # Crear DataFrame
            df = pd.DataFrame(datos_excel)
            
            # Crear el escritor de Excel
            with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as writer:
                # Escribir datos en una sola hoja
                df.to_excel(writer, sheet_name='Datos de Procesos', index=False)
            
            # Aplicar formato
            self._aplicar_formato_excel(ruta_archivo)
            
            logger.info(f"Archivo Excel creado exitosamente: {ruta_archivo}")
            return ruta_archivo
            
        except Exception as e:
            error_msg = f"Error al escribir resultados en Excel: {e}"
            logger.error(error_msg)
            raise FileManagerError(error_msg)
    
    def _aplicar_formato_excel(self, ruta_archivo: Path):
        """Aplica formato básico al archivo Excel"""
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Cargar el workbook
            wb = load_workbook(ruta_archivo)
            ws = wb.active
            
            # Estilos
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            data_font = Font(name='Calibri', size=10)
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Formatear headers (fila 1)
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            # Formatear datos
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = border
                    
                    # Alineación según columna
                    if col <= 2:  # No. y Radicado - centrado
                        cell.alignment = center_alignment
                    else:  # Resto - izquierda
                        cell.alignment = left_alignment
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 8,   # No.
                'B': 25,  # Radicado
                'C': 30,  # Demandante
                'D': 30,  # Demandado
                'E': 40,  # Juzgado
                'F': 15,  # Departamento
                'G': 20,  # Tipo del Proceso
                'H': 25,  # Clase del Proceso
                'I': 25,  # Subclase del Proceso
                'J': 12,  # Última Fecha
                'K': 50,  # Última Actuación
                'L': 60,  # Anotaciones
                'M': 10,  # Es Privado
                'N': 12   # Estado
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Congelar primera fila
            ws.freeze_panes = 'A2'
            
            # Guardar cambios
            wb.save(ruta_archivo)
            
        except Exception as e:
            logger.warning(f"No se pudo aplicar formato al Excel: {e}")
            # El archivo Excel básico ya fue creado, solo falló el formato


class FileManager:
    """Gestor principal de archivos - Simplificado para solo Excel"""
    
    def __init__(self, excel_path: Path = None, output_dir: Path = None):
        """
        Inicializa el gestor de archivos
        
        Args:
            excel_path: Ruta al archivo Excel (opcional)
            output_dir: Directorio de salida (opcional)
        """
        self.excel_reader = ExcelReader(excel_path)
        self.result_writer = ResultWriter(output_dir)
        logger.info("FileManager inicializado")
    
    def procesar_archivo_completo(self, procesos: List[ProcesoInfo], 
                                estadisticas: EstadisticasProcesamiento) -> Dict[str, Any]:
        """
        Procesa un archivo completo: lee radicados y escribe Excel
        
        Args:
            procesos: Lista de procesos procesados
            estadisticas: Estadísticas del procesamiento
            
        Returns:
            Diccionario con información del procesamiento
        """
        try:
            # Escribir solo Excel
            archivo_excel = self.result_writer.escribir_resultados_excel(procesos)
            
            # Información del archivo de entrada
            info_entrada = self.excel_reader.obtener_info_archivo()
            
            resultado = {
                "archivos_creados": {"xlsx": archivo_excel},
                "archivo_entrada": info_entrada,
                "estadisticas": {
                    "total_procesados": estadisticas.total_procesados,
                    "exitosos": estadisticas.exitosos,
                    "privados": estadisticas.privados,
                    "fallidos": estadisticas.fallidos,
                    "tasa_exito": estadisticas.tasa_exito
                }
            }
            
            logger.info("Procesamiento completo de archivo finalizado")
            return resultado
            
        except Exception as e:
            error_msg = f"Error en procesamiento completo: {e}"
            logger.error(error_msg)
            raise FileManagerError(error_msg)
    
    def validar_configuracion(self) -> Tuple[bool, List[str]]:
        """
        Valida que la configuración de archivos sea correcta
        
        Returns:
            Tupla con (es_valido, lista_errores)
        """
        errores = []
        
        # Validar archivo de entrada
        if not self.excel_reader.validar_archivo():
            errores.append(f"Archivo Excel inválido: {self.excel_reader.archivo_path}")
        
        # Validar directorio de salida
        try:
            self.result_writer.output_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            errores.append(f"No se puede crear directorio de salida: {e}")
        
        # Verificar permisos de escritura
        test_file = self.result_writer.output_dir / "test_write.tmp"
        try:
            test_file.write_text("test")
            test_file.unlink()
        except Exception as e:
            errores.append(f"Sin permisos de escritura en directorio de salida: {e}")
        
        return len(errores) == 0, errores
    
    def obtener_resumen_configuracion(self) -> Dict[str, Any]:
        """
        Obtiene un resumen de la configuración actual
        
        Returns:
            Diccionario con información de configuración
        """
        return {
            "archivo_excel": {
                "ruta": str(self.excel_reader.archivo_path),
                "existe": self.excel_reader.archivo_path.exists(),
                "info": self.excel_reader.obtener_info_archivo()
            },
            "directorio_salida": {
                "ruta": str(self.result_writer.output_dir),
                "existe": self.result_writer.output_dir.exists(),
                "archivos_existentes": len(list(self.result_writer.output_dir.glob("*"))) if self.result_writer.output_dir.exists() else 0
            },
            "configuracion": {
                "columna_excel": FileConfig.EXCEL_COLUMN,
                "fila_inicio": FileConfig.EXCEL_START_ROW,
                "encoding_salida": FileConfig.OUTPUT_ENCODING
            }
        }


class BackupManager:
    """Gestor de respaldos para archivos importantes"""
    
    def __init__(self, backup_dir: Path = None):
        """
        Inicializa el gestor de respaldos
        
        Args:
            backup_dir: Directorio de respaldos (opcional)
        """
        self.backup_dir = backup_dir or (FileConfig.PROJECT_ROOT / "backups")
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"BackupManager inicializado: {self.backup_dir}")
    
    def crear_backup_excel(self, archivo_excel: Path = None) -> Path:
        """
        Crea un respaldo del archivo Excel
        
        Args:
            archivo_excel: Ruta al archivo Excel (opcional)
            
        Returns:
            Ruta del archivo de respaldo creado
        """
        archivo_origen = archivo_excel or FileConfig.EXCEL_INPUT_FILE
        
        try:
            if not archivo_origen.exists():
                raise FileManagerError(f"Archivo a respaldar no existe: {archivo_origen}")
            
            timestamp = datetime.now().strftime(FileConfig.OUTPUT_DATETIME_FORMAT)
            nombre_backup = f"PROCESOS_backup_{timestamp}.xlsx"
            ruta_backup = self.backup_dir / nombre_backup
            
            # Copiar archivo
            shutil.copy2(archivo_origen, ruta_backup)
            
            logger.info(f"Backup creado: {ruta_backup}")
            return ruta_backup
            
        except Exception as e:
            error_msg = f"Error al crear backup: {e}"
            logger.error(error_msg)
            raise FileManagerError(error_msg)
    
    def limpiar_backups_antiguos(self, dias_antiguedad: int = 30) -> int:
        """
        Limpia backups más antiguos que el número de días especificado
        
        Args:
            dias_antiguedad: Días de antigüedad para considerar un backup como antiguo
            
        Returns:
            Número de archivos eliminados
        """
        try:
            archivos_eliminados = 0
            tiempo_limite = time.time() - (dias_antiguedad * 24 * 60 * 60)
            
            for archivo in self.backup_dir.glob("*_backup_*.xlsx"):
                if archivo.stat().st_mtime < tiempo_limite:
                    archivo.unlink()
                    archivos_eliminados += 1
                    logger.debug(f"Backup antiguo eliminado: {archivo}")
            
            if archivos_eliminados > 0:
                logger.info(f"Se eliminaron {archivos_eliminados} backups antiguos")
            
            return archivos_eliminados
            
        except Exception as e:
            logger.error(f"Error al limpiar backups antiguos: {e}")
            return 0
    
    def listar_backups(self) -> List[Dict[str, Any]]:
        """
        Lista todos los backups disponibles
        
        Returns:
            Lista de diccionarios con información de backups
        """
        backups = []
        
        try:
            for archivo in self.backup_dir.glob("*_backup_*.xlsx"):
                stats = archivo.stat()
                backups.append({
                    "nombre": archivo.name,
                    "ruta": str(archivo),
                    "tamaño_bytes": stats.st_size,
                    "fecha_creacion": datetime.fromtimestamp(stats.st_ctime).isoformat(),
                    "fecha_modificacion": datetime.fromtimestamp(stats.st_mtime).isoformat()
                })
            
            # Ordenar por fecha de creación (más reciente primero)
            backups.sort(key=lambda x: x["fecha_creacion"], reverse=True)
            
        except Exception as e:
            logger.error(f"Error al listar backups: {e}")
        
        return backups


class LogFileManager:
    """Gestor específico para archivos de log"""
    
    def __init__(self, log_dir: Path = None):
        """
        Inicializa el gestor de logs
        
        Args:
            log_dir: Directorio de logs (opcional)
        """
        self.log_dir = log_dir or (FileConfig.PROJECT_ROOT / "logs")
        self.log_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"LogFileManager inicializado: {self.log_dir}")
    
    def configurar_file_handler(self, nivel_log: int = logging.INFO) -> logging.FileHandler:
        """
        Configura un file handler para logging
        
        Args:
            nivel_log: Nivel de logging
            
        Returns:
            FileHandler configurado
        """
        timestamp = datetime.now().strftime("%Y%m%d")
        log_file = self.log_dir / f"consulta_procesos_{timestamp}.log"
        
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(nivel_log)
        
        formatter = logging.Formatter(
            '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        file_handler.setFormatter(formatter)
        
        logger.info(f"File handler configurado: {log_file}")
        return file_handler
    
    def limpiar_logs_antiguos(self, dias_antiguedad: int = 7) -> int:
        """
        Limpia logs más antiguos que el número de días especificado
        
        Args:
            dias_antiguedad: Días de antigüedad
            
        Returns:
            Número de archivos eliminados
        """
        try:
            archivos_eliminados = 0
            tiempo_limite = time.time() - (dias_antiguedad * 24 * 60 * 60)
            
            for archivo in self.log_dir.glob("*.log"):
                if archivo.stat().st_mtime < tiempo_limite:
                    archivo.unlink()
                    archivos_eliminados += 1
            
            if archivos_eliminados > 0:
                logger.info(f"Se eliminaron {archivos_eliminados} logs antiguos")
            
            return archivos_eliminados
            
        except Exception as e:
            logger.error(f"Error al limpiar logs antiguos: {e}")
            return 0


# Funciones de utilidad
def verificar_espacio_disco(directorio: Path, espacio_minimo_mb: int = 100) -> bool:
    """
    Verifica que hay suficiente espacio en disco
    
    Args:
        directorio: Directorio a verificar
        espacio_minimo_mb: Espacio mínimo requerido en MB
        
    Returns:
        True si hay suficiente espacio
    """
    try:
        _, _, espacio_libre = shutil.disk_usage(directorio)
        espacio_libre_mb = espacio_libre // (1024 * 1024)
        
        return espacio_libre_mb >= espacio_minimo_mb
        
    except Exception as e:
        logger.error(f"Error al verificar espacio en disco: {e}")
        return True  # Asumir que hay espacio si no se puede verificar


def limpiar_nombre_archivo(nombre: str) -> str:
    """
    Limpia un nombre de archivo removiendo caracteres inválidos
    
    Args:
        nombre: Nombre original
        
    Returns:
        Nombre limpio válido para sistema de archivos
    """
    import re
    
    # Remover caracteres inválidos para nombres de archivo
    caracteres_invalidos = r'[<>:"/\\|?*]'
    nombre_limpio = re.sub(caracteres_invalidos, '_', nombre)
    
    # Limitar longitud
    if len(nombre_limpio) > 200:
        nombre_limpio = nombre_limpio[:200]
    
    return nombre_limpio.strip()